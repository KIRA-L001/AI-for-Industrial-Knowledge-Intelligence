"""Document text extraction (OCR + layout) with per-format handlers.

A single `extract` entrypoint dispatches on content type. Heavy/native
dependencies (pypdf, python-docx, openpyxl, pytesseract, Pillow) are imported
lazily so the module loads and the pure-text formats (CSV/TXT) work without
them — keeping unit tests fast and dependency-free.
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from typing import Any


@dataclass
class ExtractedPage:
    """One page of extracted content."""

    page_number: int
    text: str
    blocks: list[dict[str, Any]] = field(default_factory=list)


class ExtractionError(RuntimeError):
    """Raised when a document cannot be parsed."""


def extract(data: bytes, content_type: str, filename: str = "") -> list[ExtractedPage]:
    """Extract text pages from raw bytes based on content type."""
    if content_type == "application/pdf":
        return _extract_pdf(data)
    if content_type == (
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    ):
        return _extract_docx(data)
    if content_type == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    ):
        return _extract_xlsx(data)
    if content_type == "text/csv":
        return _extract_csv(data)
    if content_type.startswith("image/"):
        return _extract_image(data)
    # Fallback: best-effort UTF-8 decode.
    return [ExtractedPage(page_number=1, text=data.decode("utf-8", errors="ignore"))]


def _extract_csv(data: bytes) -> list[ExtractedPage]:
    text_io = io.StringIO(data.decode("utf-8", errors="ignore"))
    rows = list(csv.reader(text_io))
    lines = ["\t".join(row) for row in rows]
    return [ExtractedPage(page_number=1, text="\n".join(lines))]


def _extract_pdf(data: bytes) -> list[ExtractedPage]:
    try:
        from pypdf import PdfReader
    except ImportError as exc:  # pragma: no cover - optional dep
        raise ExtractionError("pypdf is required to parse PDFs.") from exc

    reader = PdfReader(io.BytesIO(data))
    pages: list[ExtractedPage] = []
    for i, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        pages.append(ExtractedPage(page_number=i, text=text))
    return pages or [ExtractedPage(page_number=1, text="")]


def _extract_docx(data: bytes) -> list[ExtractedPage]:
    try:
        from docx import Document as DocxDocument
    except ImportError as exc:  # pragma: no cover - optional dep
        raise ExtractionError("python-docx is required to parse DOCX.") from exc

    doc = DocxDocument(io.BytesIO(data))
    text = "\n".join(p.text for p in doc.paragraphs)
    return [ExtractedPage(page_number=1, text=text)]


def _extract_xlsx(data: bytes) -> list[ExtractedPage]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - optional dep
        raise ExtractionError("openpyxl is required to parse XLSX.") from exc

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    pages: list[ExtractedPage] = []
    for idx, sheet in enumerate(wb.worksheets, start=1):
        lines = [
            "\t".join("" if c is None else str(c) for c in row)
            for row in sheet.iter_rows(values_only=True)
        ]
        pages.append(ExtractedPage(page_number=idx, text="\n".join(lines)))
    return pages or [ExtractedPage(page_number=1, text="")]


def _extract_image(data: bytes) -> list[ExtractedPage]:
    try:
        import pytesseract
        from PIL import Image
    except ImportError as exc:  # pragma: no cover - optional dep
        raise ExtractionError("pytesseract and Pillow are required for image OCR.") from exc

    image = Image.open(io.BytesIO(data))
    # Word-level boxes give bounding boxes for citation provenance.
    ocr = pytesseract.image_to_data(image, output_type=pytesseract.Output.DICT)
    words: list[str] = []
    blocks: list[dict[str, Any]] = []
    for i, word in enumerate(ocr["text"]):
        if not word.strip():
            continue
        words.append(word)
        blocks.append(
            {
                "text": word,
                "bbox": [ocr["left"][i], ocr["top"][i], ocr["width"][i], ocr["height"][i]],
                "type": "word",
            }
        )
    return [ExtractedPage(page_number=1, text=" ".join(words), blocks=blocks)]
